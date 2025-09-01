# excel_data_reader.py
# Module to read execution details and test data from the Excel file in config folder

import os
import openpyxl
import pandas as pd
import re
from modules.automation_process import process_step
from modules.reporting_v2 import RobustReporting
from modules.selenium_actions import SeleniumActions
import datetime
import shutil
import sys
import tempfile
from modules.globals import global_dict

step_results = []
CONFIG_FOLDER = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config')
EXCEL_FILE = os.path.join(CONFIG_FOLDER, 'testcase_driver_data_sheet.xlsx')

def parse_datasheet_refs(ref_str):
    """
    Parses a string like 'DataSheet!B4:DataSheet!B6' and returns (sheet_name, [row_numbers]).
    Supports single cell or range. Returns (sheet_name, [row_numbers]).
    """
    if not ref_str:
        return None, []
    # Example: DataSheet!B4:DataSheet!B6
    refs = ref_str.split(":")
    row_nums = []
    sheet_name = None
    for ref in refs:
        m = re.match(r"([\w\s]+)!([A-Z]+)(\d+)", ref)
        if m:
            sheet_name = m.group(1)
            row_nums.append(int(m.group(3)))
    if len(row_nums) == 2:
        # Range, fill in-between
        row_nums = list(range(row_nums[0], row_nums[1]+1))
    return sheet_name, row_nums

def get_temp_excel_copy():
    """Copy the Excel file to a temporary file and return its path."""
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    tmp.close()
    shutil.copy2(EXCEL_FILE, tmp.name)
    return tmp.name

def get_testcase_to_datarefs_dict(sheet_name='DriverSheet'):
    """
    Returns a dict: {TestCaseName: [(sheet_name, [row_numbers]), ...], ...}
    Only includes testcases where the first occurrence of TestCaseName has Execute == 'Y'.
    Skips blank TestDataSheetReference.
    """
    temp_excel = get_temp_excel_copy()
    wb = openpyxl.load_workbook(temp_excel, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in Excel file.")
    driver_sheet = wb[sheet_name]
    headers = [cell.value for cell in next(driver_sheet.iter_rows(min_row=1, max_row=1))]
    col_idx = {h: i for i, h in enumerate(headers)}
    if 'TestCaseName' not in col_idx or 'TestDataSheetReference' not in col_idx or 'Execute' not in col_idx:
        raise ValueError("Required columns not found in DriverSheet.")
    testcase_first_execute = {}
    testcase_datarefs = {}
    for row in driver_sheet.iter_rows(min_row=2, values_only=True):
        testcase = str(row[col_idx['TestCaseName']]).strip() if row[col_idx['TestCaseName']] else ''
        dataref = str(row[col_idx['TestDataSheetReference']]).strip() if row[col_idx['TestDataSheetReference']] else ''
        execute = str(row[col_idx['Execute']]).strip().upper() if row[col_idx['Execute']] else ''
        if not testcase or not dataref:
            continue
        # Only process testcases where the first occurrence has Execute == 'Y'
        if testcase not in testcase_first_execute:
            testcase_first_execute[testcase] = execute
            if execute != 'Y':
                continue  # Skip this testcase entirely if first occurrence is not Y
        if testcase_first_execute[testcase] != 'Y':
            continue  # Skip all subsequent rows for this testcase if first was not Y
        if testcase not in testcase_datarefs:
            testcase_datarefs[testcase] = []
        sheet, rows = parse_datasheet_refs(dataref)
        if sheet and rows:
            testcase_datarefs[testcase].append((sheet, rows))
    filtered = {tc: refs for tc, refs in testcase_datarefs.items() if testcase_first_execute.get(tc) == 'Y'}
    wb.close()
    os.unlink(temp_excel)
    return filtered

def get_component_steps(component_name, components_sheet):
    """
    Given a component name and the openpyxl worksheet for Components,
    return a list of dicts for each step in the component.
    Handles merged cells by forward-filling the ComponentName.
    """
    steps = []
    headers = [cell.value for cell in next(components_sheet.iter_rows(min_row=1, max_row=1))]
    last_component = None
    for row in components_sheet.iter_rows(min_row=2, values_only=True):
        comp_val = row[0]
        if comp_val is not None and str(comp_val).strip() != '':
            last_component = str(comp_val).strip()
        # Forward-fill merged/empty cells
        if last_component == str(component_name).strip():
            step = {headers[i]: row[i] for i in range(len(headers))}
            step[headers[0]] = last_component  # Ensure ComponentName is set
            steps.append(step)
    return steps

def process_testcase_rows(testcase, sheet, row_num, driver, reporting, actions, dataset_number):
    try:
        print(f"Processing rows for TestCase: {testcase}")
        temp_excel = get_temp_excel_copy()
        wb = openpyxl.load_workbook(temp_excel, data_only=True, read_only=True)
        driver_sheet = wb['DriverSheet']
        common_sheet = wb['CommonSheet']
        data_sheet = wb[sheet]
        components_sheet = wb['Components'] if 'Components' in wb.sheetnames else None
        # Check if the Execute column in the data_sheet for the given row_num is 'Y'
        header_row = [cell.value for cell in data_sheet[1]]
        execute_col_idx = None
        for idx, col_name in enumerate(header_row):
            if str(col_name).strip().lower() == 'execute':
                execute_col_idx = idx + 1  # 1-based index for openpyxl
                break
        if execute_col_idx:
            execute_value = data_sheet.cell(row=row_num, column=execute_col_idx).value
            if str(execute_value).strip().upper() != 'Y':
                msg = f"Skipping TestCase '{testcase}' DataSheet '{sheet}' Row {row_num}: Execute column is not 'Y'"
                reporting.log_info(msg)
                wb.close()
                return
            
        # Find GetPassScreenshot column index in DriverSheet
        driver_headers = [cell.value for cell in next(driver_sheet.iter_rows(min_row=1, max_row=1))]
        get_pass_screenshot_idx = None
        for idx, col_name in enumerate(driver_headers):
            if str(col_name).strip().lower() == 'getpassscreenshot':
                get_pass_screenshot_idx = idx
                break
        
        # Build CommonSheet lookup: {(Screen, Field): (Xpath, WaitTimeInSec)}
        common_lookup = {}
        for row in common_sheet.iter_rows(min_row=2, values_only=True):
            screen, field, xpath = row[:3]
            wait_time = 0
            if len(row) > 3 and row[3] is not None:
                try:
                    wait_time = float(row[3])
                except Exception:
                    wait_time = 0
            if screen and field and xpath:
                common_lookup[(str(screen).strip(), str(field).strip())] = (str(xpath).strip(), wait_time)
        # Find column indices in DriverSheet
        headers = [cell.value for cell in next(driver_sheet.iter_rows(min_row=1, max_row=1))]
        col_idx = {h: i for i, h in enumerate(headers)}
        # Loop through DriverSheet rows for this testcase and Execute=Y
        #steps = []
        for row in driver_sheet.iter_rows(min_row=2, values_only=True):
            tc = str(row[col_idx['TestCaseName']]).strip() if row[col_idx['TestCaseName']] else ''
            execute = str(row[col_idx['Execute']]).strip().upper() if row[col_idx['Execute']] else ''
            if tc != testcase or execute != 'Y':
                continue
            screen = str(row[col_idx['Screen']]).strip() if row[col_idx['Screen']] else ''
            field = str(row[col_idx['Field']]).strip() if row[col_idx['Field']] else ''
            action = str(row[col_idx['Action']]).strip() if row[col_idx['Action']] else ''
            testcase_description = str(row[col_idx['TestCaseDescription']]).strip() if 'TestCaseDescription' in col_idx and row[col_idx['TestCaseDescription']] else ''
            validation = str(row[col_idx['Validation']]).strip() if 'Validation' in col_idx and row[col_idx['Validation']] else ''
            expected_validation = str(row[col_idx['ExpectedValidation']]).strip() if 'ExpectedValidation' in col_idx and row[col_idx['ExpectedValidation']] else ''
            component_name = str(row[col_idx['ComponentName']]).strip() if 'ComponentName' in col_idx and row[col_idx['ComponentName']] else ''
            # Get xpath and wait_time from common_lookup
            xpath, wait_time = common_lookup.get((screen, field), ('', 0))
            data_value = None
            header_row = [cell.value for cell in data_sheet[1]]
            if field in header_row:
                col_idx_val = header_row.index(field)
                data_value = str(data_sheet.cell(row=row_num, column=col_idx_val + 1).value)
                if data_value.find("$$"):
                    key_name = data_value
                    if key_name in global_dict:
                        data_value = global_dict[key_name]
            else:
                data_value = ''

            if "<<" in xpath and ">>" in xpath:
                # Replace all occurrences of <<key>> with corresponding value from data_sheet row
                matches = re.findall(r"<<([^<>]+)>>", xpath)
                for key in matches:
                    if key in header_row:
                        col_idx_val = header_row.index(key)
                        keys_value = data_sheet.cell(row=row_num, column=col_idx_val + 1).value
                        xpath = xpath.replace(f"<<{key}>>", str(keys_value))
            
            get_pass_screenshot = False
            if get_pass_screenshot_idx is not None:
                gps_val = row[get_pass_screenshot_idx]
                if str(gps_val).strip().upper() == 'Y':
                    get_pass_screenshot = True
            # If ComponentName is present, expand steps from Components sheet
            if component_name and components_sheet:
                component_steps = get_component_steps(component_name, components_sheet)
                for comp_step in component_steps:
                    comp_screen = str(comp_step.get('Screen', '')).strip()
                    comp_field = str(comp_step.get('Field', '')).strip()
                    comp_action = str(comp_step.get('Action', '')).strip()
                    comp_testcase_description = str(comp_step.get('TestCaseDescription', '')).strip()
                    comp_validation = str(comp_step.get('Validation', '')).strip()
                    comp_expected_validation = str(comp_step.get('ExpectedValidation', '')).strip()
                    # Get xpath and wait_time from common_lookup for component step
                    comp_xpath, comp_wait_time = common_lookup.get((comp_screen, comp_field), ('', 1))
                    if not comp_wait_time:
                        comp_wait_time = 1
                    comp_data_value = ''
                    if comp_field in header_row:
                        col_idx_val = header_row.index(comp_field)
                        comp_data_value = str(data_sheet.cell(row=row_num, column=col_idx_val + 1).value)
                        if comp_data_value.find("$$"):
                            key_name = comp_data_value
                            if key_name in global_dict:
                                comp_data_value = global_dict[key_name]

                    if "<<" in comp_xpath and ">>" in comp_xpath:
                        # Replace all occurrences of <<key>> with corresponding value from data_sheet row
                        matches = re.findall(r"<<([^<>]+)>>", comp_xpath)
                        for key in matches:
                            if key in header_row:
                                col_idx_val = header_row.index(key)
                                keys_value = data_sheet.cell(row=row_num, column=col_idx_val + 1).value
                                comp_xpath = comp_xpath.replace(f"<<{key}>>", str(keys_value))

                    step_result = process_step(
                        testcase, comp_screen, comp_field, comp_action, comp_xpath, comp_data_value, driver, reporting, actions, dataset_number,
                        get_pass_screenshot=get_pass_screenshot,
                        testcase_description=comp_testcase_description,
                        validation=comp_validation,
                        expected_validation=comp_expected_validation,
                        wait_time_before_exec=comp_wait_time
                    )
                    step_result['testcase_description'] = comp_testcase_description
                    step_result['validation'] = comp_validation
                    step_result['expected_validation'] = comp_expected_validation
                    step_results.append(step_result)
            else:
                # Pass wait_time to process_step (add argument if needed)
                step_result = process_step(
                    testcase, screen, field, action, xpath, data_value, driver, reporting, actions, dataset_number,
                    get_pass_screenshot=get_pass_screenshot,
                    testcase_description=testcase_description,
                    validation=validation,
                    expected_validation=expected_validation,
                    wait_time_before_exec=wait_time
                )
                step_result['testcase_description'] = testcase_description
                step_result['validation'] = validation
                step_result['expected_validation'] = expected_validation
                step_results.append(step_result)
        wb.close()
        os.unlink(temp_excel)
    except Exception as e:
        reporting.log_error(f"Error in process_testcase_rows for testcase '{testcase}': {e}")

def initiatedriver(browser='edge', headless=True):
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.service import Service as ChromeService
        from selenium.webdriver.edge.service import Service as EdgeService
        from selenium.webdriver.chrome.options import Options as ChromeOptions
        from selenium.webdriver.edge.options import Options as EdgeOptions
        driver = None
        if browser == 'chrome':
            chrome_options = ChromeOptions()
            chrome_options.add_argument('--start-maximized')
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-gpu')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
            if headless:
                chrome_options.add_argument('--headless=new')
            driver = webdriver.Chrome(service=ChromeService(), options=chrome_options)
        elif browser == 'edge':
            edge_options = EdgeOptions()
            edge_options.add_argument('--start-maximized')
            edge_options.add_argument('--no-sandbox')
            edge_options.add_argument('--disable-gpu')
            edge_options.add_argument('--disable-dev-shm-usage')
            edge_options.add_experimental_option('excludeSwitches', ['enable-logging'])
            if headless:
                edge_options.add_argument('--headless=new')
            driver = webdriver.Edge(service=EdgeService(), options=edge_options)
        else:
            raise ValueError('Unsupported browser')
        return driver
    except Exception as e:
        RobustReporting().log_error(f"Error in initiatedriver: {e}")
        raise

# Example usage (for testing):
if __name__ == "__main__":                    
    try:
        # Create timestamped report folder
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        report_folder = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'Reports', f'execution_report_{timestamp}')
        os.makedirs(report_folder, exist_ok=True)    
        step_results = []
        # Set environment variable for screenshot saving
        os.environ['CURRENT_REPORT_FOLDER'] = report_folder

        print("\nTestCase to DataSheet row mapping:")
        testcase_datarefs = get_testcase_to_datarefs_dict()
        reporting = RobustReporting()       
        for testcase, refs in testcase_datarefs.items():
            dataset_counter = 1
            for sheet, row_nums in refs:
                print(f"  TestCase: {testcase}, DataSheet: {sheet}, Rows: {row_nums}")
                for row_num in row_nums:
                    print(f"    Row {row_num}")
                    browser = 'chrome'  # or 'edge'
                    try:
                        driver = initiatedriver(browser)
                        actions = SeleniumActions(reporting, driver)
                        print(f"TestCase: {testcase}")
                        process_testcase_rows(testcase, sheet, row_num, driver, reporting, actions, dataset_counter)
                        driver.quit()
                    except Exception as e:
                        reporting.log_error(f"Driver or step error for testcase '{testcase}', row {row_num}: {e}")
                    dataset_counter += 1
        # Write report to Excel
        report_df = pd.DataFrame(step_results)
        
        report_path = os.path.join(report_folder, f'execution_report_{timestamp}.xlsx')
        report_df.to_excel(report_path, index=False)
        print(f"Execution report saved to: {report_path}")

        # Generate HTML report
        html_template_path = os.path.join(CONFIG_FOLDER, 'html_template.html')
        html_report_path = os.path.join(report_folder, f'execution_report_{timestamp}.html')
        reporting.generate_html_report(report_path, html_template_path, html_report_path)
        print(f"Execution HTML report saved to: {html_report_path}")
    except Exception as e:
        RobustReporting().log_error(f"Fatal error in __main__: {e}")
    finally:
        if 'CURRENT_REPORT_FOLDER' in os.environ:
            del os.environ['CURRENT_REPORT_FOLDER']
