# main.py
"""
Entry point for the automation framework. This script starts the execution by calling the main function from excel_data_reader.
"""

import os
import datetime
import pandas as pd
from modules.excel_data_reader import get_testcase_to_datarefs_dict, process_testcase_rows, initiatedriver, EXCEL_FILE, CONFIG_FOLDER, step_results
from modules.reporting_v2 import RobustReporting
from modules.selenium_actions import SeleniumActions

def main():
    # Create timestamped report folder
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    report_folder = os.path.join(os.path.dirname(__file__), 'Reports', f'execution_report_{timestamp}')
    os.makedirs(report_folder, exist_ok=True)
    # Set environment variable for screenshot saving
    os.environ['CURRENT_REPORT_FOLDER'] = report_folder

    print("\nTestCase to DataSheet row mapping:")
    testcase_datarefs = get_testcase_to_datarefs_dict()
    reporting = RobustReporting()
    for testcase, refs in testcase_datarefs.items():
        dataset_counter = 1
        for sheet, row_nums in refs:
            print(f"  TestCase: {testcase}, DataSheet: {sheet}, Rows: {row_nums}")
            for row_num in row_nums:
                print(f"    Row {row_num}")
                # Determine headless mode from DataSheet's 'HeadLess' column for this row
                headless = True  # Default to headless
                try:
                    import openpyxl
                    datasheet_path = os.path.join(CONFIG_FOLDER, 'testcase_driver_data_sheet.xlsx')
                    wb = openpyxl.load_workbook(datasheet_path, data_only=True)
                    if sheet in wb.sheetnames:
                        data_sheet = wb[sheet]
                        header_row = [cell.value for cell in data_sheet[1]]
                        if 'HeadLess' in header_row:
                            col_idx = header_row.index('HeadLess') + 1  # 1-based
                            headless_cell = data_sheet.cell(row=row_num, column=col_idx).value
                            if str(headless_cell).strip().upper() != 'T':
                                headless = False
                    wb.close()
                except Exception as e:
                    print(f"[WARN] Could not determine headless mode for {sheet} row {row_num}: {e}")
                browser = 'edge'  # Use Edge as the default browser
                driver = initiatedriver(browser, headless=headless)
                actions = SeleniumActions(reporting, driver)
                print(f"TestCase: {testcase}")
                process_testcase_rows(testcase, sheet, row_num, driver, reporting, actions, dataset_counter)
                driver.quit()
                dataset_counter += 1
    # Write report to Excel
    report_df = pd.DataFrame(step_results)
    report_path = os.path.join(report_folder, f'execution_report_{timestamp}.xlsx')
    report_df.to_excel(report_path, index=False)
    print(f"Execution report saved to: {report_path}")

    # Generate HTML report
    html_template_path = os.path.join(CONFIG_FOLDER, 'html_template.html')
    html_report_path = os.path.join(report_folder, f'execution_report_{timestamp}.html')
    reporting.generate_html_report(report_path, html_template_path, html_report_path)
    print(f"Execution HTML report saved to: {html_report_path}")

    # Clean up environment variable
    if 'CURRENT_REPORT_FOLDER' in os.environ:
        del os.environ['CURRENT_REPORT_FOLDER']

if __name__ == "__main__":
    main()
